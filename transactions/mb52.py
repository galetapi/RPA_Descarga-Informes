import sys
import win32com.client as win32
from datetime import datetime
from openpyxl import load_workbook
#from app import routes


def mb52(key,log):

  try:

    SapGuiAuto = win32.GetObject("SAPGUI")
    if not type(SapGuiAuto) == win32.CDispatch:
      return

    application = SapGuiAuto.GetScriptingEngine
    if not type(application) == win32.CDispatch:
      SapGuiAuto = None
      return

    connection = application.Children(0)
    if not type(connection) == win32.CDispatch:
      application = None
      SapGuiAuto = None
      return

    session = connection.Children(0)
    if not type(session) == win32.CDispatch:

      connection = None
      application = None
      SapGuiAuto = None
      return

    ##Ruta para lectura excel##
    file_path = key
    sheet = 'Hoja1'
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook[sheet]

    ##Lectura de campos del excel
    transaction = sheet.cell(row=2,column=6).value
    variant = sheet.cell(row=3,column=6).value
    disposition = sheet.cell(row=4,column=6).value
    route = sheet.cell(row=1,column=2).value
    nameTransaction = "/N" + transaction
    errorMail = sheet.cell(row=1,column=6).value

    #Nombre del archivo (fecha y hora actual)    
    date = datetime.today().strftime("%Y-%m-%d")
    nameFile = str(date + transaction)
    ##Elementos mapeo
    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/tbar[0]/okcd").text = nameTransaction
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/tbar[1]/btn[17]").press()
    session.findById("wnd[1]/usr/txtV-LOW").text = variant
    
    session.findById("wnd[1]/usr/txtENAME-LOW").text = ""
    session.findById("wnd[1]/usr/txtV-LOW").caretPosition = 13
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    session.findById("wnd[0]/usr/ctxtP_VARI").text = disposition
    barraTexto = session.findById("wnd[0]/sbar").text


    if len(barraTexto) > 0 :
          outlook = win32.Dispatch('outlook.application')
          mail = outlook.CreateItem(0)
          mail.To = errorMail
          mail.Subject = "No hay cargue de datos"
          mail.Body = "La transaccion no arrojo datos"
          mail.Send()
          
    session.findById("wnd[0]/usr/ctxtP_VARI").setFocus()
    session.findById("wnd[0]/usr/ctxtP_VARI").caretPosition = 12
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    session.findById("wnd[0]/tbar[1]/btn[45]").press()
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").select()
    session.findById("wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]").setFocus()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = route

    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = nameFile
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 3
    session.findById("wnd[1]/tbar[0]/btn[0]").press() 
    session.findById("wnd[0]/tbar[0]/btn[12]").press()
    session.findById("wnd[0]/tbar[0]/btn[12]").press()
  
  except:
    print(sys.exc_info([0]))
    log.Level='ERROR'
    log.ProcesoInterno='MB52'
    log.Mensaje='Variante no encontrada'
    log.post()

  finally:
    session = None
    connection = None
    application = None
    SapGuiAuto = None
