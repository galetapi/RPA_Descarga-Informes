import sys
import win32com.client as win32
from datetime import datetime
from openpyxl import load_workbook


def tablero(key,log):

  try:

    SapGuiAuto = win32.GetObject("SAPGUI")
    if not type(SapGuiAuto) == win32.CDispatch:
      return

    application = SapGuiAuto.GetScriptingEngine
    if not type(application)== win32.CDispatch:
      SapGuiAuto = None
      return

    connection = application.Children(0)
    if not type(connection) == win32.CDispatch:
      application = None
      SapGuiAuto = None
      return

    session = connection.Children(0)
    if not type(session) == win32.CDispatch:

      connection = None
      application = None
      SapGuiAuto = None
      return

    file_path = key
    sheet = 'Hoja1'
    workbook = load_workbook(file_path)
    sheet = workbook[sheet]

    transaction = sheet.cell(row=2,column=4).value
    variant = sheet.cell(row=3,column=4).value
    dispotition = sheet.cell(row=4, column=4).value
    route = sheet.cell(row=1,column=2).value
    nameTransaction = "/N" + transaction
    errorMail = sheet.cell(row=1,column=6).value
  
    date = datetime.today().strftime("%Y-%m-%d")
    nameFile = str(date +" "+ transaction)

    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/tbar[0]/okcd").text = nameTransaction
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/tbar[1]/btn[17]").press()
    session.findById("wnd[1]/usr/txtV-LOW").text = variant    
    session.findById("wnd[1]/usr/txtENAME-LOW").text = ""
    session.findById("wnd[1]/usr/txtV-LOW").caretPosition = 4
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    session.findById("wnd[0]/usr/shell").pressToolbarButton ("&MB_VARIANT")
    row = 1
    for i in range(100):         
      x = session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").GetCellValue(i, "VARIANT")
      if x == dispotition:
          row = i
          break

    
    barraTexto = session.findById("wnd[0]/sbar").text


    if len(barraTexto) > 0 :
          outlook = win32.Dispatch('outlook.application')
          mail = outlook.CreateItem(0)
          mail.To = errorMail
          mail.Subject = "No hay cargue de datos"
          mail.Body = "La transaccion no arrojo datos"
          mail.Send()

    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").currentCellRow = row
    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").selectedRows = row
    session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").clickCurrentCell()    
    session.findById("wnd[0]/usr/shell").setCurrentCell (3,"MAKTX")
    session.findById("wnd[0]/usr/shell").contextMenu()
    session.findById("wnd[0]/usr/shell").selectContextMenuItem ("&XXL")
    session.findById("wnd[1]/usr/radRB_OTHERS").setFocus()
    session.findById("wnd[1]/usr/radRB_OTHERS").select()
    session.findById("wnd[1]/usr/cmbG_LISTBOX").setFocus()
    session.findById("wnd[1]/usr/cmbG_LISTBOX").key = "10"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = route
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = nameFile+ ".xls"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 22
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[0]/tbar[0]/btn[12]").press()
    session.findById("wnd[0]/tbar[0]/btn[12]").press()

  except:
    print(sys.exc_info([0]))
    log.level = 'ERROR'
    log.procesointerno = 'ZPP_TABLERO'
    log.mensaje = 'Variante no encontrada'
    log.post()


  finally:
    session = None
    connection = None
    application = None
    SapGuiAuto = None
